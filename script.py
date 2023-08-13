#from cgitb import reset
from datetime import datetime
import re
import os
import openpyxl

# Function to reset Flags
def reset_flags():
    global capture_next_row, captured_next_row, skip_next_row
    capture_next_row = False
    captured_next_row = False
    skip_next_row = True

# Function to reset Index
def reset_index():
    global row_index, val_index 
    row_index = 2
    val_index = 0

# Function to convert date from 'YYYYMMDD' to 'DD/MM/YYYY' format
def convert_date_format(date_str):
    date_obj = datetime.strptime(date_str, '%Y%m%d')
    return date_obj.strftime('%d/%m/%Y')

# Function to extract F35B data from a line
def extract_f35b_data(line):
    pattern = r"\S+(?=\s)"
    matches = re.findall(pattern, line)
    return matches[-1] if matches else None

# Function to extract F36B data from a line
def extract_f36b_data(line):
    pattern = r"\b\d+"
    matches = re.findall(pattern, line)
    return matches[0] if matches else None

# Function to extract F19A: Posting Amount  data from a line
def extract_f19a_posting_data(line):
    pattern = r"[\d,]+"
    matches = re.findall(pattern, line)
    result = "".join(matches)
    return matches[0] if matches else None

# Function to extract F20C: Account Owner Reference data from a line
def extract_f20c_account_data(line):
    pattern = r"\d+"
    matches = re.findall(pattern, line)
    return matches[0] if matches else None

# Function to extract F24B: Pending/Failing Reason data from a line
def extract_f24b_pending_data(line):
    pattern = r"Reason Code:\s*\/(\w+)"
    matches = re.findall(pattern, line)
    return matches[0] if matches else None

# Function to extract F19A: Securities Side data from a line
def extract_f19a_securities_data(line):
    pattern = r"[\d,]+"
    matches = re.findall(pattern, line)
    result = "".join(matches)
    return matches[0] if matches else None

def extract_f92a_corporate_bonds_data(line):
    pattern = r"[\d,]+"
    matches = re.findall(pattern, line)
    result = "".join(matches)
    return matches[0] if matches else None

# Function to extract F97A: Account Number data from a line
def extract_f97a_account_data(line):
    pattern = r"\d+"
    matches = re.findall(pattern, line)
    return matches[0] if matches else None
    
# Function to extract F23G data from a line
def extract_f23g_data(line):
    pattern = r"\b[A-Z]{4}\b"
    matches = re.findall(pattern, line)
    return matches[0] if matches else None

# Function to extract F20C (SEME) data from a line
def extract_f20c_seme_data(line):
    pattern = r"SEME\s*/(\w+)\s*"
    match = re.search(pattern, line)
    return match.group(1) if match else None

# Function to extract F20C: Penalty Common Reference data from a line
def extract_f20c_penalty_data(line):
    pattern = r"PCOM\s*/(\w+)\s*"
    match = re.search(pattern, line)
    return match.group(1) if match else None

# Function to extract F20C: Penalty Reference data from a line
def extract_f20c_penalty_reference_data(line):
    pattern = r"PREF\s*/(\w+)\s*"
    match = re.search(pattern, line)
    return match.group(1) if match else None

# Function to extract F95R: Related Party - Data Source Scheme - Proprietary Code data from a line
def extract_f95r_prop_code_data(line):
    pattern = r"Proprietary Code:\s*(\S+)\s*"
    match = re.search(pattern, line)
    return match.group(1) if match else None

# Assume the files are in the local drive in the "files" directory
directory_path = "F:/project/bca/MT537 Converter/attachments/"  # Replace this with the actual path to your files directory

# Get a list of all files in the directory
all_files = os.listdir(directory_path)

# Filter the list to include only .txt files
txt_files = [file for file in all_files if file.endswith(".txt")]

#Inisialisasi Variabel global pertama kali
skip_next_row = True  # Flag to skip the next row after FXXX found
capture_next_row = False
captured_next_row = False  # Variable to track if next row data has been captured
row_index = 2
val_index = 0

# Process each .txt file one by one
for txt_file in txt_files:
    file_path = os.path.join(directory_path, txt_file)

    with open(file_path, "r") as file:
        lines = file.readlines()

    # Process F98A data
    f98a_row_val= ""
    pattern = r"STAT\s*/(\w+)\s*"
    for line in lines:
        if 'F98A: Date' in line:
            capture_next_row = True

        elif capture_next_row and not captured_next_row :
            f98a_row_val = re.search(pattern, line) # Extract the date part
            f98a_row_val = f98a_row_val.group(1)
            capture_next_row = False
            #formatted_date = f98a_row_val.replace('/', '')
            formatted_date = f98a_row_val
            formatted_date = convert_date_format(formatted_date)
            reset_flags()
            break
    
    # Process F69A data
    data_dict_f69a = {}
    for line in lines:
        if 'F69A:' in line:
            header = "F69A"
            capture_next_row = True
        elif capture_next_row:
            if 'Date1:' in line:
                date_value1 = line.split("Date1:")[1].replace("/", "").strip()[:8]
                date_formatted1 = convert_date_format(date_value1)
                data_dict_f69a[f"{header}: Period - Date1 TRADE DATE"] = date_formatted1
            elif 'Date2:' in line:
                date_value2 = line.split("Date2:")[1].replace("/", "").strip()[:8]
                date_formatted2 = convert_date_format(date_value2)
                data_dict_f69a[f"{header}: Period - Date2 SETTLE DATE"] = date_formatted2
                capture_next_row = False
                skip_next_row = False
                reset_flags()

    # Process F35B data
    f35b_row_val = ""
    # Append all value found in file
    f35_val_list =[]

    for line in lines:
        if 'F35B:' in line:
            capture_next_row = True
    
        elif capture_next_row and not captured_next_row :
            f35b_row_val = extract_f35b_data(line)  # Extract the date part from F35B line
            f35_val_list.append(f35b_row_val)
            reset_flags()
            #break
    
    # Process F36B data
    f36b_row_val = ""
    # Append all value found in file
    f36b_val_list =[]
    for line in lines:
        if 'F36B:' in line:
            capture_next_row = True
            
        elif capture_next_row and not captured_next_row and 'Quantity:' in line :
            f36b_row_val = extract_f36b_data(line)  # Extract the date part from F35B line
            f36b_val_list.append(f36b_row_val)
            reset_flags()
            #break
    
    # Process F19A: Posting Amount  data
    f19a_posting_row_val = ""
    # Append all value found in file
    f19a_val_list =[]
    for line in lines:
        if 'F19A: Posting Amount' in line:
            capture_next_row = True
            
        elif capture_next_row and not captured_next_row and 'Amount:' in line :
            f19a_posting_row_val =extract_f19a_posting_data(line)  # Extract the date part from F35B line
            f19a_val_list.append(f19a_posting_row_val)
            reset_flags()
            #break

    # Process F20C: Account Owner Reference data
    f20c_account_row_val = ""
    # Append all value found in file
    f20c_account_row_val_list =[]
    for line in lines:
        if 'F20C: Account Owner Reference' in line:
            capture_next_row = True
            
        elif capture_next_row and not captured_next_row and 'ACOW' in line :
            f20c_account_row_val = extract_f20c_account_data(line)  # Extract the date part from F35B line
            f20c_account_row_val_list.append(f20c_account_row_val)
            reset_flags()
            #break

    # Process F24B: Pending/Failing Reason data
    f24b_pending_row_val = ""
    # Append all value found in file
    f24b_val_list =[]
    for line in lines:
        if 'F24B: Pending' in line:
            capture_next_row = True
            
        elif capture_next_row and not captured_next_row and 'Reason' in line :
            f24b_pending_row_val = extract_f24b_pending_data(line)  # Extract the date part from F35B line
            f24b_val_list.append(f24b_pending_row_val)
            reset_flags()
            #break
    
    # Process F19A: Securities Side data
    f19a_securities_row_val = ""
    # Append all value found in file
    f19a_securities_val_list =[]
    for line in lines:
        if 'F19A: Securities Side' in line:
            capture_next_row = True

        elif capture_next_row and not captured_next_row and 'Amount' in line:
            f19a_securities_row_val = extract_f19a_securities_data(line)  # Extract the data from F19A: Securities Side line
            f19a_securities_val_list.append(f19a_securities_row_val)
            reset_flags()
            #break

    # Process F92A: Corporate Bonds data
    f92a_corporate_bonds_row_val = ""
    # Append all value found in file
    f92a_corporate_bonds_val_list =[]
    for line in lines:
        if 'F92A:' in line:
            capture_next_row = True
            
        elif capture_next_row and not captured_next_row and 'Rate:' in line:
            f92a_corporate_bonds_row_val = extract_f92a_corporate_bonds_data(line)
            f92a_corporate_bonds_val_list.append(f92a_corporate_bonds_row_val)
            reset_flags()
            #break
    
    # Process F97A: Account Number data
    f97a_account_row_val = ""
    # Append all value found in file
    f97a_account_val_list =[]
    for line in lines:
        if 'F97A: Account Number' in line:
            capture_next_row = True
        
        elif 'SAFE /NONREF' in line:
            continue
            
        elif capture_next_row and not captured_next_row and 'SAFE' in line:
            f97a_account_row_val = extract_f97a_account_data(line)  # Extract the data part from F97A line
            f97a_account_val_list.append(f97a_account_row_val)
            reset_flags()
            #break

    # Process F23G: Function of the Message data
    f23g_row_val = ""
    for line in lines:
        if 'F23G:' in line:
            capture_next_row = True

        elif capture_next_row and not captured_next_row and 'Function:' in line:
            f23g_row_val = extract_f23g_data(line)  # Extract F23G: Function of the Message data
            reset_flags()
            break

    # Process F20C (SEME) data
    f20c_seme_row_val = ""
    for line in lines:
        if 'F20C (SEME):' in line:
            capture_next_row = True
        
        elif capture_next_row and not captured_next_row and 'SEME' in line:
            f20c_seme_row_val = extract_f20c_seme_data(line)  # Extract F20C (SEME) data 
            reset_flags()
            break
    
    # Process F20C: Penalty Common Reference data
    f20c_penalty_row_val = ""
    # Append all value found in file
    f20c_penalty_val_list =[]
    for line in lines:
        if 'F20C: Penalty Common Reference' in line:
            capture_next_row = True
        elif capture_next_row and not captured_next_row and 'PCOM' in line:
            f20c_penalty_row_val = extract_f20c_penalty_data(line)  # Extract F20C: Penalty Common Reference data
            f20c_penalty_val_list.append(f20c_penalty_row_val)
            reset_flags()
            #break
    
    # Process F20C: Penalty Reference data
    f20c_penalty_reference_row_val = ""
    # Append all value found in file
    f20c_penalty_reference_val_list =[]
    for line in lines:
        if 'F20C: Penalty Reference' in line:
            capture_next_row = True
        elif capture_next_row and not captured_next_row and 'PREF' in line:
            f20c_penalty_reference_row_val = extract_f20c_penalty_reference_data(line)  # Extract F20C: Penalty Reference data
            f20c_penalty_reference_val_list.append(f20c_penalty_reference_row_val)
            reset_flags()
            #break
    
    # Process F95R: Related Party - Data Source Scheme - Proprietary Code data
    f95r_prop_code_row_val = ""
    # Append all value found in file
    f95c_prop_code_val_list =[]
    for line in lines:
        if 'F95R: Related Party - Data Source Scheme - Proprietary Code' in line:
            capture_next_row = True
        elif capture_next_row and not captured_next_row and 'Proprietary Code' in line:
        #elif capture_next_row and not captured_next_row and len(f95c_prop_code_val_list) < len(f20c_penalty_reference_val_list)  and 'Proprietary Code' in line:
            f95r_prop_code_row_val = extract_f95r_prop_code_data(line)  # Extract F95R: Related Party - Data Source Scheme - Proprietary Code data
            f95c_prop_code_val_list.append(f95r_prop_code_row_val)
            reset_flags()
            #break
    
    # Save the data to an Excel file
    wb = openpyxl.Workbook()
    ws = wb.active

    #formatted_date = f98a_row_val.replace('/', '')
    #formatted_date = f98a_row_val
    #formatted_date = convert_date_format(formatted_date)

    # Add F98A data to Excel file
    if f98a_row_val:
    
        if len(f35_val_list) > 1 :
            for index,item in enumerate (f35_val_list,0):
                ws.cell(row=row_index, column=1, value=formatted_date)
                row_index+=1
                val_index+=1
        else:
            ws.cell(row=2, column=1, value=formatted_date)
        
        ws.cell(row=1, column=1, value="F98A: Date")
        col_width = max(len("F98A: Date"), len(formatted_date))
        ws.column_dimensions['A'].width = col_width
        reset_index()

    # Add F69A data to Excel file
    if len(f35_val_list) > 1 :
        for index,item in enumerate (f35_val_list,0):
            ws.cell(row=row_index, column=2, value=data_dict_f69a.get("F69A: Period - Date1 TRADE DATE", ""))
            ws.cell(row=row_index, column=3, value=data_dict_f69a.get("F69A: Period - Date2 SETTLE DATE", ""))
            row_index+=1
            val_index+=1
    else :
        ws.cell(row=2, column=2, value=data_dict_f69a.get("F69A: Period - Date1 TRADE DATE", ""))
        ws.cell(row=2, column=3, value=data_dict_f69a.get("F69A: Period - Date2 SETTLE DATE", ""))

    ws.cell(row=1, column=2, value="F69A: Period - Date1 TRADE DATE")
    ws.cell(row=1, column=3, value="F69A: Period - Date2 SETTLE DATE")
    col_width_f69a_col2 = max(len("F69A: Period - Date1 TRADE DATE"), max(len(date) for date in data_dict_f69a.values()))
    ws.column_dimensions['B'].width = col_width_f69a_col2
    ws.column_dimensions['C'].width = col_width_f69a_col2
    reset_index()

    # Add F35B data to Excel file
    if f35b_row_val:
        ws.cell(row=1, column=4, value="F35B: Identification of the Financial Instrument")
        for index,item in enumerate (f35_val_list,0):
            ws.cell(row=row_index, column=4, value=f35_val_list[val_index])
            col_width_f35b_col = max(len("F35B: Identification of the Financial Instrument"), len(f35b_row_val))
            ws.column_dimensions['D'].width = col_width_f35b_col
            row_index+=1
            val_index+=1
        reset_index()

    # Add F36B data to Excel file
    if f36b_row_val:
        ws.cell(row=1, column=5, value="F36B: Posting Quantity - Quantity Type Code - Quantity")
        for index,item in enumerate (f36b_val_list,0):
            ws.cell(row=row_index, column=5, value=f36b_val_list[val_index])
            col_width_f36b_col = max(len("F36B: Posting Quantity - Quantity Type Code - Quantity"), len(f36b_row_val))
            ws.column_dimensions['E'].width = col_width_f36b_col    
            row_index+=1
            val_index+=1
        reset_index()

    # Add F19A: Posting Amount data to Excel file
    if f19a_posting_row_val:
        ws.cell(row=1, column=6, value="F19A: Posting Amount")
        for index,item in enumerate (f19a_val_list,0):
            ws.cell(row=row_index, column=6, value=f19a_val_list[val_index])
            col_width_f19b_posting_col = max(len("F19A: Posting Amount"), len(f19a_posting_row_val))
            ws.column_dimensions['F'].width = col_width_f19b_posting_col
            row_index+=1
            val_index+=1
        reset_index() 

    # Add F20C: Account Owner Reference data to Excel file
    if f20c_account_row_val:
        ws.cell(row=1, column=7, value="F20C: Account Owner Reference")
        for index,item in enumerate (f20c_account_row_val_list,0):
            ws.cell(row=row_index, column=7, value=f20c_account_row_val_list[val_index])
            col_width_f20_account_col = max(len("F20C: Account Owner Reference"), len(f20c_account_row_val))
            ws.column_dimensions['G'].width = col_width_f20_account_col 
            row_index+=1
            val_index+=1
        reset_index()

    # Add F24B: Pending/Failing Reason data to Excel file
    if f24b_pending_row_val:
        ws.cell(row=1, column=8, value="F24B: Pending/Failing Reason")
        for index,item in enumerate (f24b_val_list,0):
            ws.cell(row=row_index, column=8, value=f24b_val_list[val_index])
            col_width_f24_pending_col = max(len("F24B: Pending/Failing Reason"), len(f24b_pending_row_val))
            ws.column_dimensions['H'].width = col_width_f24_pending_col
            row_index+=1
            val_index+=1
        reset_index()

    # Add F19A: Securities Side data to Excel file
    if f19a_securities_row_val:
        ws.cell(row=1, column=9, value="F19A: Securities Side (amount)")
        for index,item in enumerate (f19a_securities_val_list,0):
            ws.cell(row=row_index, column=9, value=f19a_securities_val_list[val_index])
            col_width_f19a_securities_col = max(len("F19A: Securities Side (amount)"), len(f19a_securities_row_val))
            ws.column_dimensions['I'].width = col_width_f19a_securities_col
            row_index+=1
            val_index+=1
        reset_index()

    # Add F92A: Corporate Bonds data to Excel file
    if f92a_corporate_bonds_row_val:
        ws.cell(row=1, column=10, value="F92A: Corporate Bonds (rate)")
        for index,item in enumerate (f92a_corporate_bonds_val_list,0):
            ws.cell(row=row_index, column=10, value=f92a_corporate_bonds_val_list[val_index])
            col_width_f92a_col = max(len("F92A: Corporate Bonds (rate)"), len(f92a_corporate_bonds_row_val))
            ws.column_dimensions['J'].width = col_width_f92a_col
            row_index+=1
            val_index+=1
        reset_index()

    # Add F97A: Account Number data to E7xcel file
    if f97a_account_row_val:
        ws.cell(row=1, column=11, value="F97A: Account Number")
        for index,item in enumerate (f97a_account_val_list,0):
            ws.cell(row=row_index, column=11, value=f97a_account_val_list[val_index])
            col_width_f97a_col = max(len("F97A: Account Number"), len(f97a_account_row_val))
            ws.column_dimensions['K'].width = col_width_f97a_col
            row_index+=1
            val_index+=1
        reset_index()

    # Add F23G data to Excel file
    if f23g_row_val:

        if len(f35_val_list) > 1 :
            for index, item in enumerate (f35_val_list,0):
                ws.cell(row=row_index, column=12, value=f23g_row_val)
                row_index+=1
                val_index+=1
        else:
            ws.cell(row=row_index, column=12, value=f23g_row_val)
        
        ws.cell(row=1, column=12, value="F23G: Function of the Message")
        col_width_f23g_col = max(len("F23G: Function of the Message"), len(f23g_row_val))
        ws.column_dimensions['L'].width = col_width_f23g_col
        reset_index()

    # Add F20C (SEME) data to Excel file
    if f20c_seme_row_val:

        if len(f35_val_list) > 1 :
            for index, item in enumerate (f35_val_list,0):
                ws.cell(row=row_index, column=13, value=f20c_seme_row_val)
                row_index+=1
                val_index+=1
        else:
            ws.cell(row=2, column=13, value=f20c_seme_row_val)
        
        ws.cell(row=1, column=13, value="F20C (SEME): Reference")
        col_width_f20c_seme_col = max(len("F20C (SEME): Reference"), len(f20c_seme_row_val))
        ws.column_dimensions['M'].width = col_width_f20c_seme_col
        reset_index()

    # Add F20C: Penalty Common Reference data to Excel file
    if f20c_penalty_row_val:
        ws.cell(row=1, column=14, value="F20C: Penalty Common Reference")
        for index, item in enumerate (f20c_penalty_val_list):
            ws.cell(row=row_index, column=14, value=f20c_penalty_val_list[val_index])
            col_width_f20c_penalty_col = max(len("F20C: Penalty Common Reference"), len(f20c_penalty_row_val))
            ws.column_dimensions['N'].width = col_width_f20c_penalty_col
            row_index+=1
            val_index+=1
        reset_index()

    # Add F20C: Penalty Reference data to Excel file
    if f20c_penalty_reference_row_val:
        ws.cell(row=1, column=15, value="F20C: Penalty Reference")
        for index,item in enumerate (f20c_penalty_reference_val_list):
            ws.cell(row=row_index, column=15, value=f20c_penalty_reference_val_list[val_index])
            col_width_f20c_penalty_reference_col = max(len("F20C: Penalty Reference"), len(f20c_penalty_reference_row_val))
            ws.column_dimensions['O'].width = col_width_f20c_penalty_reference_col
            row_index+=1
            val_index+=1
        reset_index()

    # Add F95R: Related Party - Data Source Scheme - Proprietary Code data to Excel file
    if f95r_prop_code_row_val:
        ws.cell(row=1, column=16, value="F95R: Related Party - Data Source Scheme - Proprietary Code (Proprietary Code)")
        for index, item in enumerate (f35_val_list):
            ws.cell(row=row_index, column=16, value=f95c_prop_code_val_list[val_index])
            col_width_f95r_prop_code_col = max(len("F95R: Related Party - Data Source Scheme - Proprietary Code (Proprietary Code)"), len(f95r_prop_code_row_val))
            ws.column_dimensions['P'].width = col_width_f95r_prop_code_col
            row_index+=1
            val_index+=1
        reset_index()

    output_file = os.path.splitext(txt_file)[0] + ".xlsx"
    wb.save(output_file)
